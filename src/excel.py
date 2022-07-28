from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill  # 导入填充模块
from OSACA import capstoneInput
from llvm_mca import capstone
import global_variable as glv
import time
from multiBar import time2String
from data import dataDictInit
from tsjPython.tsjCommonFunc import *
from KendallIndex import calculateKendallIndex

def addData2Excel(wb,taskName,isFirstSheet,dataDict):
    [llvmerror,baselineError,validBlockNum,validInstructionNum] = add2Excel(wb,taskName,isFirstSheet,dataDict)
    if glv._get("KendallIndex")=="yes":
        [KendallIndex, baselineKendallIndex]=calculateKendallIndex(dataDict)
    elif glv._get("KendallIndex")=="no":
        KendallIndex=0
        baselineKendallIndex=0
    excelGraphAdd(wb,taskName,llvmerror,baselineError,validBlockNum,validInstructionNum,KendallIndex, baselineKendallIndex)

# https://blog.csdn.net/David_Dai_1108/article/details/78702032
def readExcelSet(pageName, columnName):
    yellowPrint("Reading excel {} data……".format(columnName))
    unique_revBiblock=set()
    if not columnName=="block_binary":
        return unique_revBiblock
    from openpyxl import load_workbook
    # 只读模式打开文件
    wb = load_workbook(glv._get("HistoryDataFile"), read_only=True)
    # 获得所有 sheet 的名称()
    name_list = wb.sheetnames
    # 根据 sheet 名字获得 sheet
    if pageName not in name_list:
        ic(pageName,"Not existed")
        return unique_revBiblock
    ic(pageName,"existed")
    my_sheet = wb[pageName]
    for rowNum in range(2,my_sheet.max_row):
        ic(rowNum)
        unique_revBiblock.add(my_sheet["B"+str(rowNum)].value)
    return unique_revBiblock

def readExcelDict(pageName,columnName):
    yellowPrint("Reading excel {} data……".format(columnName))
    readDefaultdict = defaultdict(int)
    if not(columnName=="BHive" or columnName=="Baseline"):
        return readDefaultdict
    elif columnName=="llvm-mca":
        columnLetter="F"
    elif columnName=="Baseline":
        columnLetter="G"
    elif columnName=="BHive":
        columnLetter="H"
    from openpyxl import load_workbook
    # 只读模式打开文件
    wb = load_workbook(glv._get("HistoryDataFile"), read_only=True)
    # 获得所有 sheet 的名称()
    name_list = wb.sheetnames
    # 根据 sheet 名字获得 sheet
    if pageName not in name_list:
        ic(pageName,"Not existed")
        return readDefaultdict
    ic(pageName,"existed")
    my_sheet = wb[pageName]
    for rowNum in range(2,my_sheet.max_row):
        ic(rowNum)
        readDefaultdict[my_sheet["B"+str(rowNum)].value]=int(my_sheet[columnLetter+str(rowNum)].value)
    return readDefaultdict

def readDictFromExcel(taskName):
    import time
    processBeginTime=time.time()
    colorPrint("\n\rstart read excel at: {}".format(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())),"magenta")
    historyDict = dataDictInit()
    if glv._get("useBhiveHistoryData")=="no" and glv._get("useBaselineHistoryData")=="no" and glv._get("useLLVMHistoryData")=="no":
        ic("hb hm lv is all NO!")
        glv._set("isPageExisted","no")
        return historyDict
    history_unique_revBiblock = readExcelSet(taskName,"block_binary")
    if not history_unique_revBiblock:
        ic(taskName,"is Not Existed!")
        glv._set("isPageExisted","no")
        return historyDict
    ic(taskName,"is Existed!")
    glv._set("isPageExisted","yes")
    history_BHive=readExcelDict(taskName,"BHive")
    history_LLVM=readExcelDict(taskName,"llvm-mca")
    history_Baseline=readExcelDict(taskName,"Baseline")
    historyDict.dataDict["unique_revBiblock"]=historyDict.dataDict["unique_revBiblock"].union(history_unique_revBiblock)
    historyDict.dataDict["BhiveCyclesRevBiBlock"].update(history_BHive)
    historyDict.dataDict["llvmmcaCyclesRevBiBlock"].update(history_LLVM)
    historyDict.dataDict["BaselineCyclesRevBiBlock"].update(history_Baseline)
    colorPrint("wait {} to finish read excel at: {}".format(time2String(int(time.time()-processBeginTime)),time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())),"magenta")
    return historyDict

def add2Excel(wb,name,isFirstSheet,dataDict):
    #unique_revBiblock,frequencyRevBiBlock,OSACAmaxCyclesRevBiBlock,OSACACPCyclesRevBiBlock,OSACALCDCyclesRevBiBlock,BhiveCyclesRevBiBlock,accuracyMax,accuracyCP,llvmmcaCyclesRevBiBlock,accuracyLLVM):
    # if isFirstSheet==1:
        # ws = wb.active # 找当前Sheet
        # ws.title = name
    wb.create_sheet(name)
    ws = wb[name]
    ws.append(["num","block_binary" , "ARM64_assembly_code", "Num_of_instructions","block_frequency","block_frequency_percentage","LLVM-MCA_result", "Baseline","BHive", "accuracyLLVM", "accuracyBaseline","accuracyLLVM * block_frequency" ,"accuracyBaseline * block_frequency" ]) # 添加行
    ws.column_dimensions['B'].width = 80 # 修改列宽
    ws.column_dimensions['K'].width = 30 # 修改列宽
    ws.column_dimensions['L'].width = 40 # 修改列宽
    ws.column_dimensions['M'].width = 40 # 修改列宽
    ws.column_dimensions['C'].width = 40 # 修改列宽
    for i in ['D','E','F','G','H','I','J','K']:
        ws.column_dimensions[i].width = 20 # 修改列宽
    validInstructionNum=0
    unvalidNum=0
    BhiveSkipNum=0
    totalAccuracyLLVM=0.0
    totalAccuracyLLVMBaseline=0.0
    # totalaccuracyMax=0.0
    # totalaccuracyCP=0.0
    # totalOSACAavg=0.0
    
    for key, value in dataDict.dataDict.items():
        globals()[key]=value 

    # 先统计validInstructionNum总数
    for tmp_block_binary_reverse in unique_revBiblock:
        if BhiveCyclesRevBiBlock[tmp_block_binary_reverse]==-1:
            continue
        if accuracyLLVM[tmp_block_binary_reverse] != 0:
            validInstructionNum+=frequencyRevBiBlock[tmp_block_binary_reverse]

    lineNum=0
    for tmp_block_binary_reverse in unique_revBiblock:
        if BhiveCyclesRevBiBlock[tmp_block_binary_reverse]==-1:
            BhiveSkipNum+=1
            continue
        lineNum+=1
        tmpARMassembly=capstone(capstoneInput(tmp_block_binary_reverse))
        ic(len(tmp_block_binary_reverse))
        if accuracyLLVM[tmp_block_binary_reverse] != 0:
            # totalAccuracyLLVM+=frequencyRevBiBlock[tmp_block_binary_reverse]*accuracyLLVM[tmp_block_binary_reverse]
            frequencyPercentage=frequencyRevBiBlock[tmp_block_binary_reverse]*1.0/validInstructionNum
            totalAccuracyLLVM+=accuracyLLVM_MuliplyFrequency[tmp_block_binary_reverse]
            totalAccuracyLLVMBaseline += accuracyBaseline_MuliplyFrequency[tmp_block_binary_reverse]
            # totalaccuracyMax+=frequencyRevBiBlock[tmp_block_binary_reverse]*accuracyMax[tmp_block_binary_reverse]
            # totalaccuracyCP+=frequencyRevBiBlock[tmp_block_binary_reverse]*accuracyCP[tmp_block_binary_reverse]

            # count=frequencyRevBiBlock[tmp_block_binary_reverse]
            # OSACAMax=OSACAmaxCyclesRevBiBlock[tmp_block_binary_reverse]
            # OSACACP=OSACACPCyclesRevBiBlock[tmp_block_binary_reverse]
            # realBHive=float(BhiveCyclesRevBiBlock[tmp_block_binary_reverse])
            # tmp=(OSACAMax+OSACACP)/2 * BHiveCount - realBHive
            # totalOSACAavg+=abs(tmp)/realBHive * count
            # 会显得表格特别稀疏
            # ws.row_dimensions[lineNum+1].height = int((len(tmp_block_binary_reverse)+1)/8) * 13.5

            ws.append(["{:5d} ".format(lineNum), 
                tmp_block_binary_reverse,
                tmpARMassembly,
                int((len(tmp_block_binary_reverse)+1)/8),
                frequencyRevBiBlock[tmp_block_binary_reverse],
                '%.2f%%' % (frequencyPercentage * 100),
                # OSACAmaxCyclesRevBiBlock[tmp_block_binary_reverse],
                # OSACACPCyclesRevBiBlock[tmp_block_binary_reverse],
                # OSACALCDCyclesRevBiBlock[tmp_block_binary_reverse],
                llvmmcaCyclesRevBiBlock[tmp_block_binary_reverse],
                BaselineCyclesRevBiBlock[tmp_block_binary_reverse],
                BhiveCyclesRevBiBlock[tmp_block_binary_reverse],
                accuracyLLVM[tmp_block_binary_reverse],
                accuracyBaseline[tmp_block_binary_reverse],
                accuracyLLVM_MuliplyFrequency[tmp_block_binary_reverse],
                accuracyBaseline_MuliplyFrequency[tmp_block_binary_reverse]]
                # accuracyMax[tmp_block_binary_reverse],
                # accuracyCP[tmp_block_binary_reverse]]
                )

            toFill=ws['F{}'.format(lineNum+1)]
            if frequencyPercentage > 0.2:
                toFill.fill=PatternFill('solid', fgColor='ffeb9c') # 黄
            elif frequencyPercentage > 0.1:
                toFill.fill=PatternFill('solid', fgColor='FFFFCC') # 浅黄

            
            if not llvmmcaCyclesRevBiBlock[tmp_block_binary_reverse]==BaselineCyclesRevBiBlock[tmp_block_binary_reverse]:
                if accuracyLLVM[tmp_block_binary_reverse]<accuracyBaseline[tmp_block_binary_reverse]:
                    toFill=ws['G{}'.format(lineNum+1)]
                    toFill.fill=PatternFill('solid', fgColor='c6efce') #绿
                else:
                    toFill=ws['H{}'.format(lineNum+1)]
                    toFill.fill=PatternFill('solid', fgColor='ffeb9c') # 黄

            toFill=ws['J{}'.format(lineNum+1)]
            if accuracyLLVM[tmp_block_binary_reverse] > 1:
                toFill.fill=PatternFill('solid', fgColor='FF0000') # 红
            elif accuracyLLVM[tmp_block_binary_reverse] > 0.5:
                toFill.fill=PatternFill('solid', fgColor='ffc7ce') # 浅红
            elif accuracyLLVM[tmp_block_binary_reverse] > 0.25:
                toFill.fill=PatternFill('solid', fgColor='FFFF00') # 黄
            elif accuracyLLVM[tmp_block_binary_reverse] > 0.1:
                toFill.fill=PatternFill('solid', fgColor='FFFF99') # 浅黄
            
            toFill=ws['K{}'.format(lineNum+1)]
            if accuracyBaseline[tmp_block_binary_reverse] > 1:
                toFill.fill=PatternFill('solid', fgColor='FF0000') # 红
            elif accuracyBaseline[tmp_block_binary_reverse] > 0.5:
                toFill.fill=PatternFill('solid', fgColor='ffc7ce') # 浅红
            elif accuracyBaseline[tmp_block_binary_reverse] > 0.25:
                toFill.fill=PatternFill('solid', fgColor='FFFF00') # 黄
            elif accuracyBaseline[tmp_block_binary_reverse] > 0.1:
                toFill.fill=PatternFill('solid', fgColor='FFFF99') # 浅黄


        else:
            unvalidNum+=1
            # ws.row_dimensions[lineNum+1].height = int((len(tmp_block_binary_reverse)+1)/8) * 13.5
            frequencyPercentage=frequencyRevBiBlock[tmp_block_binary_reverse]*1.0/validInstructionNum
            ws.append(["{:5d} ".format(lineNum), 
                tmp_block_binary_reverse,
                tmpARMassembly,
                int((len(tmp_block_binary_reverse)+1)/8),
                frequencyRevBiBlock[tmp_block_binary_reverse],
                '%.2f%%' % (frequencyPercentage * 100),
                # OSACAmaxCyclesRevBiBlock[tmp_block_binary_reverse],
                # OSACACPCyclesRevBiBlock[tmp_block_binary_reverse],
                # OSACALCDCyclesRevBiBlock[tmp_block_binary_reverse],
                llvmmcaCyclesRevBiBlock[tmp_block_binary_reverse],
                BaselineCyclesRevBiBlock[tmp_block_binary_reverse],
                BhiveCyclesRevBiBlock[tmp_block_binary_reverse],
                accuracyLLVM[tmp_block_binary_reverse],
                accuracyBaseline[tmp_block_binary_reverse],
                accuracyLLVM_MuliplyFrequency[tmp_block_binary_reverse],
                accuracyBaseline_MuliplyFrequency[tmp_block_binary_reverse],
                # accuracyMax[tmp_block_binary_reverse],
                # accuracyCP[tmp_block_binary_reverse]]
                "ops!"])
    ws.append(["validTotalBlockNum(allow duplicates)  {:d}".format(validInstructionNum),"llvmUnvalidNum {:d}".format(unvalidNum),"BhiveSkipNum {:d}".format(BhiveSkipNum)])
    if validInstructionNum==0:
        llvmerror=0
        baselineError=0
    else:
        llvmerror=totalAccuracyLLVM/validInstructionNum
        baselineError=totalAccuracyLLVMBaseline/validInstructionNum
    # osacaerror=totalOSACAavg/validInstructionNum
    # osacaerror = 0
    return [llvmerror,baselineError,lineNum,validInstructionNum]

def excelGraphInit():
    wb = Workbook()
    ws = wb.active # 找当前Sheet
    ws.title = 'Graph'
    ws = wb["Graph"]
    ws.column_dimensions['A'].width = 15 # 修改列宽
    ws.column_dimensions['B'].width = 15 # 修改列宽
    ws.column_dimensions['C'].width = 15 # 修改列宽
    ws.column_dimensions['D'].width = 15 # 修改列宽
    ws.column_dimensions['E'].width = 15 # 修改列宽
    ws.column_dimensions['F'].width = 15 # 修改列宽
    ws.column_dimensions['G'].width = 15 # 修改列宽
    ws.column_dimensions['H'].width = 15 # 修改列宽
    ws.append(["applications","LLVM-MCA_error","baseline_error",'误差比值','有效Block数','指令总数(包括重复的)',"KendallIndex", "baselineKendallIndex"])
    return wb

def excelGraphAdd(wb,taskName,llvmerror,baselineError,validBlockNum,validInstructionNum,KendallIndex, baselineKendallIndex):
    ws = wb["Graph"]
    ws.append([taskName,llvmerror,baselineError,0,validBlockNum,validInstructionNum,KendallIndex, baselineKendallIndex])
    wb.save(glv._get("excelOutPath"))

def excelGraphBuild(wb,processBeginTime):
    # 一个图两个轴
    ws = wb["Graph"]
    ct_bar = BarChart()
    taskNum=len(glv._get("taskList"))
    ws['D1'] = '误差比值'
    for i in range(2, taskNum+2):
        toFill=ws[f'B{i}']
        toFillValue=ws[f'B{i}'].value
        if toFillValue > 0.5:
            toFill.fill=PatternFill('solid', fgColor='FF0000') #红
        elif toFillValue > 0.2:
            toFill.fill=PatternFill('solid', fgColor='ffc7ce') #浅红
        elif toFillValue > 0.1:
            toFill.fill=PatternFill('solid', fgColor='FFFF00') #黄
        elif toFillValue <= 0.1:
            toFill.fill=PatternFill('solid', fgColor='c6efce') #绿
        toFill=ws[f'C{i}']
        toFillValue=ws[f'C{i}'].value
        if toFillValue > 0.5:
            toFill.fill=PatternFill('solid', fgColor='FF0000') #红
        elif toFillValue > 0.2:
            toFill.fill=PatternFill('solid', fgColor='ffc7ce') #浅红
        elif toFillValue > 0.1:
            toFill.fill=PatternFill('solid', fgColor='FFFF00') #黄
        elif toFillValue <= 0.1:
            toFill.fill=PatternFill('solid', fgColor='c6efce') #绿


        if ws[f'B{i}'].value==0:
            ws[f'D{i}']=0
        else:
            ratio = ws[f'C{i}'].value / ws[f'B{i}'].value
            ws[f'D{i}'] = ratio
            toFill=ws[f'D{i}']
            if ratio > 2:
                toFill.fill=PatternFill('solid', fgColor='c6efce') #绿
            elif ratio < 1:
                toFill.fill=PatternFill('solid', fgColor='ffc7ce') #红

            # https://openpyxl.readthedocs.io/en/stable/styles.html
    d_ref = Reference(ws, min_col=2, min_row=1, max_row=taskNum+1, max_col=3)
    ct_bar.add_data(d_ref, titles_from_data=True)
    series = Reference(ws, min_col=1, min_row=2, max_row=taskNum+1)
    ct_bar.set_categories(series)
    ct_bar.x_axis.title = '应用'
    ct_bar.y_axis.title = '误差'
    ct_bar.y_axis.majorGridlines = None
    ct_bar.title = '各应用静态分析误差对比表'
    ct_line = LineChart()
    d_ref = Reference(ws, min_col=4, min_row=1, max_row=taskNum+1)
    ct_line.add_data(d_ref, titles_from_data=True)
    ct_line.y_axis.axId = 200 # 不为空即可
    ct_line.y_axis.title = '误差比值'
    # 让线条和第一图的最大值相交
    ct_line.y_axis.crosses = 'max'
    ct_bar += ct_line # 只支持+=赋值，不能直接+
    ws.add_chart(ct_bar, 'A30')
    ws.append(["time spent {}".format(time2String(int(time.time()-processBeginTime)))])
    excelAddHeatmap(ws)
    wb.save(glv._get("excelOutPath"))

def excelAddHeatmap(ws):
    position=1
    taskList = glv._get("taskList")
    for taskKey, taskName in taskList.items():
        img = Image("./pictures/"+taskName+".png")
        img.width, img.height=(300,3*90)
        ws.add_image(img, 'K'+str(position))
        img = Image("./pictures/"+taskName+"_baseline.png")
        img.width, img.height=(300,3*90)
        ws.add_image(img, 'P'+str(position))
        position+=15