from collections import defaultdict
from multiprocessing import Queue
from logPrint import *
import sys
import global_variable as glv
from terminal_command import checkFile

class dataDictClass():
    def __init__(self):
        """在主模块初始化"""
        self.dataDict = {}

    def set(self, name, value):
        """设置"""
        try:
            self.dataDict[name] = value
            return True
        except KeyError:
            return False

    def get(self, name):
        """取值"""
        try:
            return self.dataDict[name]
        except KeyError:
            return "Not Found"

def dataDictInit():
    dataDict = dataDictClass()

    dataDict.set("unique_revBiblock",set())
    dataDict.set("frequencyRevBiBlock" , defaultdict(int))
    dataDict.set("llvmmcaCyclesRevBiBlock" , defaultdict(int))
    dataDict.set("BaselineCyclesRevBiBlock" , defaultdict(int))
    dataDict.set("OSACA_TPLCDmax_CyclesRevBiBlock" , defaultdict(float))
    dataDict.set("OSACA_TPLCDavg_CyclesRevBiBlock" , defaultdict(float))
    dataDict.set("OSACATPCyclesRevBiBlock" , defaultdict(float))
    dataDict.set("OSACALCDCyclesRevBiBlock" , defaultdict(float))
    dataDict.set("BhiveCyclesRevBiBlock" , defaultdict(int))
    dataDict.set("accuracyLLVM" , defaultdict(float))
    dataDict.set("accuracyLLVM_MuliplyFrequency" , defaultdict(float))
    dataDict.set("accuracyBaseline" , defaultdict(float))
    dataDict.set("accuracyBaseline_MuliplyFrequency" , defaultdict(float))
    dataDict.set("accuracyMax" , defaultdict(float))
    dataDict.set("accuracyAvg" , defaultdict(float))
    dataDict.set("accuracyCP" , defaultdict(float))
    return dataDict

def queueDictInit(dataDict):
    queueDict = dataDictClass()

    for key in dataDict.dataDict:
        ic(key)
        queueDict.set(key, Queue())
    return queueDict

def isExcelPageExisted(pageName):
    from openpyxl import load_workbook
    # 只读模式打开文件
    wb = load_workbook(glv._get("HistoryDataFile"), read_only=True)
    # 获得所有 sheet 的名称()
    name_list = wb.sheetnames
    # 根据 sheet 名字获得 sheet
    if pageName not in name_list:
        ic(pageName,"Not existed")
        return False
    else:
        return True
        
def readDictFromJson(taskName):
    historyDict = dataDictInit()
    if glv._get("useBhiveHistoryData")=="no" and glv._get("useBaselineHistoryData")=="no" and glv._get("useLLVMHistoryData")=="no" and glv._get("useOSACAHistoryData")=="no":
        ic("hb hm lv is all NO!")
        glv._set("isPageExisted","no")
        return historyDict
    if not isExcelPageExisted(taskName):
        ic(taskName,"is Not Existed!")
        glv._set("isPageExisted","no")
        return historyDict
    ic(taskName,"is Existed!")
    glv._set("isPageExisted","yes")
    beginTime=timeBeginPrint(sys._getframe().f_code.co_name)
    import json
    checkFile(glv._get("HistoryDataFile")+"_data/")
    rf = open(glv._get("HistoryDataFile")+"_data/"+taskName)      # 将文件中的数据反序列化成内置的dict类型
    # rf = open(glv._get("HistoryDataFile")+"_"+taskName)      # 将文件中的数据反序列化成内置的dict类型
    historyDict.dataDict=json.load(fp=rf)
    timeEndPrint(sys._getframe().f_code.co_name,beginTime)
    return historyDict

def set_default(obj):
    if isinstance(obj, set):
        return list(obj)
    raise TypeError
def saveDict2Json(taskName,dataDict):
    beginTime=timeBeginPrint(sys._getframe().f_code.co_name)
    import json
    checkFile(glv._get("excelOutPath")+"_data/")
    wf = open(glv._get("excelOutPath")+"_data/"+taskName, 'w')  # 将dict类型对象序列化存储到文件中           
    json.dump(obj=dataDict, fp=wf, default=set_default)
    timeEndPrint(sys._getframe().f_code.co_name,beginTime)
    