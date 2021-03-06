
from collections import defaultdict
import re
import os
from rich.progress import track
from rich.progress import Progress
from tqdm import *
import  time
import sys
from capstone import *
from multiprocessing import Process, Queue
import pathlib
import datetime
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, LineChart, Reference

taskfilePath="/home/shaojiemike/blockFrequency"
# taskList={\
    # "tensorflow_test_100":"tensorflow_1",
            # "tensorflow_test_5":"tensorflow_2",
            # "tensorflow_test_3":"tensorflow_3",
taskList={           "test_insns_blockFrequency_skip_2":"test_insns"}
# taskList={ "test_insns_test_5":"test"}
# taskList={ 
taskList={        
            "test_insns_blockFrequency_skip_2":"test_insns",
            "clang_harness_00all_skip_2":"Clang",
            "tensorflow_41Gdir_00all_skip_2":"Tensorflow",
            "MM_median_all_skip_2":"Eigen",
            "Gzip_all_skip_2":"Gzip",
            "redis_r1000000_n2000000_P16_all_skip_2":"Redis"}
excelOutPath = taskfilePath+'/Summary.xlsx'
# OSACAPath="/home/shaojiemike/github/OSACA-feature-tsv110/newOSACA/bin/osaca "
OSACAPath="/home/qcjiang/softwares/anaconda3/bin/osaca"
LLVM_mcaPath="/home/qcjiang/codes/llvm-project/build/bin/llvm-mca"
BHivePath="/home/qcjiang/codes/KunpengWorkload/micro_benchmarks/bhive-reg/main"
saveInfo="0222newOSACAagain"
BHiveCount=100
ProcessNum=30

def OSACA(password,inputFile,maxOrCP):
    val=os.popen('echo '+password+' | python '+OSACAPath+' --arch TSV110 '+str(inputFile))#  -timeline -show-encoding -all-stats -all-views
    list = val.readlines()
    if list:
        lineNum=1
        while lineNum:
            listObj=list[lineNum]
            regexResult=re.search("Loop-Carried Dependencies Analysis Report",listObj)
            if regexResult:
                break
            lineNum+=1
        resultLineNum=lineNum-5
        it=re.finditer("[.0-9]+",list[resultLineNum])
        resultList=[]
        try:
            if not it:
                return -1
            for match in it: 
                resultList.append(float("0"+match.group()))
            if len(resultList)>2:
                LCD=resultList.pop()
                CP=resultList.pop()
                Max=max(resultList)
                if maxOrCP == "max":
                    return Max
                elif maxOrCP == "CP":
                    return CP
                elif maxOrCP == "LCD":
                    return LCD
                else:
                    return -1
            else:
                return -1
        except Exception as e:
            print("osacaText:{}\n".format(list[resultLineNum]))
            pprint.pprint(list)
            raise e
        # for match in it: 
        #     resultList.append(float("0" + match.group()))
        # if resultList:
        #     LCD=resultList.pop()
        #     CP=resultList.pop()
        #     Max=max(resultList)
        #     if maxOrCP == "max":
        #         return Max
        #     elif maxOrCP == "CP":
        #         return CP
        #     elif maxOrCP == "LCD":
        #         return LCD
        #     else:
        #         return -1
        # else:
        #     return -1

    else:
        return -1

def saveOSACAInput2File(InputAsmList,rank):
    writeFilename="{}/tmpOSACAfiles/{}.{}_OSACAInputTmpAsmBlockRank{}".format(taskfilePath,taskfilenameprefixWithoutPath,saveInfo,rank)
    fwriteblockfreq = open(writeFilename, "w")
    for tmp_InputAsmList in InputAsmList:
        fwriteblockfreq.writelines(tmp_InputAsmList)
    fwriteblockfreq.close()
    return writeFilename

def capstoneList(string):
    CODE = bytes.fromhex(string)
    md = Cs(CS_ARCH_ARM64, CS_MODE_LITTLE_ENDIAN)
    InputAsmList=[]
    for i in md.disasm(CODE, 0x1000):
        # print("%s\t%s" %(i.mnemonic, i.op_str))
        InputAsmList.append("{}\t{}\n".format(i.mnemonic, i.op_str))
    return InputAsmList

def capstoneInput(block):
    input2word=""
    for i in range(int((len(block)+1)/9)):
        # print(i)
        input2word+=block[i*9:i*9+2]+block[i*9+2:i*9+4]
        input2word+=block[i*9+4:i*9+6]+block[i*9+6:i*9+8]
    return input2word

def LLVM_mca(password,input):
    sys.stdout.flush()
    val=os.popen('echo "'+input+'" | sudo '+LLVM_mcaPath+' -iterations='+str(BHiveCount))#  -timeline -show-encoding -all-stats -all-views
    list = val.readlines()
    #Total Cycles:      10005
    # print(list[2])
    regexResult=re.search("Total Cycles:      ([0-9]*)",list[2])
    if regexResult:
        resultCycle=regexResult.group(1)
        return resultCycle
    else:
        # print("wrong")
        return -1

def capstone(string):
    CODE = bytes.fromhex(string)
    md = Cs(CS_ARCH_ARM64, CS_MODE_LITTLE_ENDIAN)
    InputAsm=""
    for i in md.disasm(CODE, 0x1000):
        # print("%s\t%s" %(i.mnemonic, i.op_str))
        InputAsm+="{}\t{}\n".format(i.mnemonic, i.op_str)
    return InputAsm

def arroundPercent(percent,A,B):
    intA=int(A)
    intB=int(B)
    avg=(intA+intB)/2
    delta=percent*avg/100
    if intA>avg-delta and intA<avg+delta and intB>avg-delta and intB<avg+delta:
        return True
    else:
        return False


def checkBHiveResultStable(password,input,showinput,trytime):
    global BHiveCount
    order=0 
    if trytime==0:
        order=order+1
    elif trytime>5:
        return -1
    sys.stdout.flush()
    val=os.popen('echo '+password+' | sudo -S '+BHivePath+' '+input)
    list = val.readlines()
    if list is None or len(list)==0:
        regexResult=None
    else:
        regexResult=re.search("Event num: ([0-9]*)",list[-1])
    if regexResult:
        resultCycle=regexResult.group(1)
        return resultCycle
    else:
        return checkBHiveResultStable(password,input,showinput,trytime+1)

def BHive(password,input,showinput,trytime):
    global BHiveCount
    order=0 
    if trytime==0:
        order=order+1
    elif trytime>5:
        # print("trytime {}/{} {}".format(order,num_file,input))
        # print("trytime {}/{} {}".format(order,num_file,showinput))
        # print("trytime > 5")
        return -1
    # print("before main {}/{} {}".format(order,num_file,input))
    # print("before main {}/{} {}".format(order,num_file,showinput))
    sys.stdout.flush()
    # begin_time=time.time()
    val=os.popen('echo '+password+' | sudo -S '+BHivePath+' '+input)
    # call_main_time=time.time()-begin_time
    # print("after  main {}/{} {}s".format(order,num_file,call_main_time))
    # sys.stdout.flush()
    list = val.readlines()
    # real_main_time=time.time()-begin_time
    # print("real  main {}/{} {}s".format(order,num_file,real_main_time))
    # if real_main_time>20:
    #     print("Attention!")
    # sys.stdout.flush()
    # print(list[-1])
    if list is None or len(list)==0:
        regexResult=None
    else:
        regexResult=re.search("Event num: ([0-9]*)",list[-1])
    if regexResult:
        resultCycle=regexResult.group(1)
        # print(resultCycle)
        return resultCycle
	#checkCycle=checkBHiveResultStable(password,input,showinput,0)
        #if arroundPercent(5,resultCycle,checkCycle):
        #    return resultCycle
        #else:
        #    return BHive(password,input,showinput,trytime+1)   
    else:
        # print("trytime: {} {}".format(trytime ,list[-1]))
        # print("else {}/{} {}".format(order,num_file,input))
        # print("else {}/{} {}".format(order,num_file,showinput))
        return BHive(password,input,showinput,trytime+1)

def BHiveInput(block):
    # print(block)
    # print((len(block)+1)/9)
    input2word=""
    for i in range(int((len(block)+1)/9)):
        # print(i)
        input2word+=" 0x"+block[i*9:i*9+2]+" 0x"+block[i*9+2:i*9+4]
        input2word+=" 0x"+block[i*9+4:i*9+6]+" 0x"+block[i*9+6:i*9+8]
    # print(input2word)
    return input2word

def BHiveInputDel0xSpace(block):
    # print(block)
    # print((len(block)+1)/9)
    input2word=""
    for i in range(int((len(block)+1)/9)):
        # print(i)
        input2word+=block[i*9:i*9+2]+block[i*9+2:i*9+4]
        input2word+=block[i*9+4:i*9+6]+block[i*9+6:i*9+8]
    # print(input2word)
    return input2word

def BHiveInputDel0x(block):
    # print(block)
    # print((len(block)+1)/9)
    input2word=""
    for i in range(int((len(block)+1)/9)):
        # print(i)
        input2word+=" "+block[i*9:i*9+2]+" "+block[i*9+2:i*9+4]
        input2word+=" "+block[i*9+4:i*9+6]+" "+block[i*9+6:i*9+8]
    # print(input2word)
    return input2word

def calculateAccuracyOSACA(accurateCycles,predictionCycles,rank):
    # print(" rank{}-tsj".format(rank))
    if int(accurateCycles) == -1 or float(predictionCycles) == -1 or int(accurateCycles) == 0:
        # print(" rank{}-0".format(rank))
        return 0
    else:
        # print(" rank{}-tsj2".format(rank))
        # print("{} {}".format(accurateCycles,predictionCycles))
        gap=abs(BHiveCount*float(predictionCycles)-int(accurateCycles))
        # print(" rank{}-{}".format(rank,int(gap)/int(accurateCycles)))
        return int(gap)/int(accurateCycles) # accuracy variable is error

def calculateAccuracyLLVM(accurateCycles,predictionCycles):
    if int(accurateCycles) == -1 or int(predictionCycles) == -1 or int(accurateCycles) == 0:
        return 0
    else:
        # print("{} {}".format(accurateCycles,predictionCycles))
        gap=abs(int(predictionCycles)-int(accurateCycles))
        return int(gap)/int(accurateCycles) # accuracy variable is error

def paralleReadProcess(rank,password, startFileLine,endFileLine,unique_revBiblock_Queue,frequencyRevBiBlock_Queue,OSACAmaxCyclesRevBiBlock_Queue,OSACACPCyclesRevBiBlock_Queue,OSACALCDCyclesRevBiBlock_Queue,BhiveCyclesRevBiBlock_Queue,accuracyMax_Queue,accuracyCP_Queue,llvmmcaCyclesRevBiBlock_Queue,accuracyLLVM_Queue):
    print("MPI Process Start {:2d} {}~{}".format(rank,startFileLine,endFileLine))
    fread=open(filename, 'r')
    unique_revBiblock=set()
    frequencyRevBiBlock = defaultdict(int)
    llvmmcaCyclesRevBiBlock = defaultdict(int)
    OSACAmaxCyclesRevBiBlock = defaultdict(int)
    OSACACPCyclesRevBiBlock = defaultdict(int)
    OSACALCDCyclesRevBiBlock =  defaultdict(int)
    BhiveCyclesRevBiBlock = defaultdict(int)
    accuracyLLVM = defaultdict(float)
    accuracyMax = defaultdict(float)
    accuracyCP = defaultdict(float)
    for line in tqdm(fread.readlines()[startFileLine:endFileLine],total=endFileLine-startFileLine,desc=str("{:2d}".format(rank))):
    # for line in fread:
        # print("rank{}".format(rank))
        block=re.search('^(.*),',line).group(1)
        num=re.search(',(.*)$',line).group(1)
        unique_revBiblock.add(block)
        frequencyRevBiBlock[block] += int(num)
        # print("     rank{}:block{}".format(rank,block))
        BhiveCyclesRevBiBlock[block] = BHive(password,BHiveInputDel0xSpace(block),BHiveInputDel0xSpace(block),0)
        # print("         rank{}:block{}______{}".format(rank,block,BhiveCyclesRevBiBlock[block]))
        llvmmcaCyclesRevBiBlock[block] = LLVM_mca(password,capstone(capstoneInput(block)))
        OSACAInput=saveOSACAInput2File(capstoneList(capstoneInput(block)),rank)
        OSACAmaxCyclesRevBiBlock[block] = OSACA(password,OSACAInput,"max")
        OSACACPCyclesRevBiBlock[block] = OSACA(password,OSACAInput,"CP")
        OSACALCDCyclesRevBiBlock[block] = OSACA(password,OSACAInput,"LCD")
        # print("             rank{}:block{}______{}".format(rank,block,OSACAmaxCyclesRevBiBlock[block]))
        accuracyLLVM[block]= calculateAccuracyLLVM(BhiveCyclesRevBiBlock[block],llvmmcaCyclesRevBiBlock[block])
        accuracyMax[block]= calculateAccuracyOSACA(BhiveCyclesRevBiBlock[block],OSACAmaxCyclesRevBiBlock[block],rank)
        accuracyCP[block]= calculateAccuracyOSACA(BhiveCyclesRevBiBlock[block],OSACACPCyclesRevBiBlock[block],rank)
        # print("0rank{}".format(rank))
    fread.close() 
    # print("1rank{}".format(rank))
    unique_revBiblock_Queue.put(unique_revBiblock)
    frequencyRevBiBlock_Queue.put(frequencyRevBiBlock)
    # print("2rank{}".format(rank))
    llvmmcaCyclesRevBiBlock_Queue.put(llvmmcaCyclesRevBiBlock)
    OSACAmaxCyclesRevBiBlock_Queue.put(OSACAmaxCyclesRevBiBlock)
    OSACACPCyclesRevBiBlock_Queue.put(OSACACPCyclesRevBiBlock)
    OSACALCDCyclesRevBiBlock_Queue.put(OSACALCDCyclesRevBiBlock)
    BhiveCyclesRevBiBlock_Queue.put(BhiveCyclesRevBiBlock)
    # print("3rank{}".format(rank))
    accuracyLLVM_Queue.put(accuracyLLVM)
    accuracyMax_Queue.put(accuracyMax)
    accuracyCP_Queue.put(accuracyCP)
    print("MPI Process end {:2d} {}~{}".format(rank,startFileLine,endFileLine))

def readPartFile(password, unique_revBiblock,frequencyRevBiBlock,OSACAmaxCyclesRevBiBlock,OSACACPCyclesRevBiBlock,OSACALCDCyclesRevBiBlock,BhiveCyclesRevBiBlock,accuracyMax,accuracyCP,llvmmcaCyclesRevBiBlock,accuracyLLVM):
    global num_file,ProcessNum
    fread=open(filename, 'r')
    num_file = sum([1 for i in open(filename, "r")])
    fread.close() 

    unique_revBiblock_Queue =Queue()
    frequencyRevBiBlock_Queue = Queue()
    llvmmcaCyclesRevBiBlock_Queue=Queue()
    OSACAmaxCyclesRevBiBlock_Queue=Queue()
    OSACACPCyclesRevBiBlock_Queue=Queue()
    OSACALCDCyclesRevBiBlock_Queue=Queue()
    BhiveCyclesRevBiBlock_Queue=Queue()
    accuracyLLVM_Queue=Queue()
    accuracyMax_Queue=Queue()
    accuracyCP_Queue=Queue()

    for i in range(ProcessNum):
        startFileLine=int(i*num_file/ProcessNum)
        endFileLine=int((i+1)*num_file/ProcessNum)
        p = Process(target=paralleReadProcess, args=(i,password, startFileLine,endFileLine,unique_revBiblock_Queue,frequencyRevBiBlock_Queue,OSACAmaxCyclesRevBiBlock_Queue,OSACACPCyclesRevBiBlock_Queue,OSACALCDCyclesRevBiBlock_Queue,BhiveCyclesRevBiBlock_Queue,accuracyMax_Queue,accuracyCP_Queue,llvmmcaCyclesRevBiBlock_Queue,accuracyLLVM_Queue))
        p.start()

    while unique_revBiblock_Queue.qsize()<ProcessNum:
        print("QueueNum : {}".format(unique_revBiblock_Queue.qsize()))
        sys.stdout.flush()
        time.sleep(5)
    # for i in tqdm(range(ProcessNum)):
    for i in range(ProcessNum):
        print("MPISum rank : {}, blockNum : {},leftQueueNum : {}".format(i,len(unique_revBiblock),unique_revBiblock_Queue.qsize()))
        unique_revBiblock=unique_revBiblock.union(unique_revBiblock_Queue.get())
        frequencyRevBiBlock.update(frequencyRevBiBlock_Queue.get())
        llvmmcaCyclesRevBiBlock.update(llvmmcaCyclesRevBiBlock_Queue.get())
        OSACAmaxCyclesRevBiBlock.update(OSACAmaxCyclesRevBiBlock_Queue.get())
        OSACACPCyclesRevBiBlock.update(OSACACPCyclesRevBiBlock_Queue.get())
        OSACALCDCyclesRevBiBlock.update(OSACALCDCyclesRevBiBlock_Queue.get())
        BhiveCyclesRevBiBlock.update(BhiveCyclesRevBiBlock_Queue.get())
        accuracyLLVM.update(accuracyLLVM_Queue.get())
        accuracyMax.update(accuracyMax_Queue.get())
        accuracyCP.update(accuracyCP_Queue.get())
    return unique_revBiblock
    # print(unique_revBiblock)
    # print(frequencyRevBiBlock)
    # print(accuracyMax)
    # print(len(unique_revBiblock))
    # print(len(frequencyRevBiBlock))
    


def saveAllResult(taskfilenameprefix,unique_revBiblock,frequencyRevBiBlock,OSACAmaxCyclesRevBiBlock,OSACACPCyclesRevBiBlock,BhiveCyclesRevBiBlock,accuracyMax,accuracyCP,llvmmcaCyclesRevBiBlock,accuracyLLVM):
    writeFilename="{}_count{}_{}_OSACAVSLLVMVSBHive.csv".format(taskfilenameprefix,BHiveCount,saveInfo)
    wrongResultFilename="{}_count{}_{}_OSACAVSLLVMVSBHive_wrongResult.csv".format(taskfilenameprefix,BHiveCount,saveInfo)
    fwriteblockfreq = open(writeFilename, "w")
    wrongResultFile = open(wrongResultFilename, "w")
    fwriteblockfreq.writelines("{:5d} ".format(0)+', block_binary, block_frequency, OSACAmax, OSACACP, LLVM-MCA, BHive, accuracyLLVM, accuracyMax, accuracyCP\n')
    wrongResultFile.writelines("{:5d} ".format(0)+', block_binary, block_frequency, OSACAmax, OSACACP, LLVM-MCA, BHive, accuracyLLVM, accuracyMax, accuracyCP\n')
    validNum=0
    unvalidNum=0
    totalAccuracyLLVM=0.0
    totalaccuracyMax=0.0
    totalaccuracyCP=0.0
    totalOSACAavg=0.0
    lineNum=0
    for tmp_block_binary_reverse in unique_revBiblock:
        lineNum+=1
        if accuracyMax[tmp_block_binary_reverse] != 0 and accuracyCP[tmp_block_binary_reverse] != 0 and accuracyLLVM[tmp_block_binary_reverse] != 0:
            validNum+=frequencyRevBiBlock[tmp_block_binary_reverse]
            totalAccuracyLLVM+=frequencyRevBiBlock[tmp_block_binary_reverse]*accuracyLLVM[tmp_block_binary_reverse]
            totalaccuracyMax+=frequencyRevBiBlock[tmp_block_binary_reverse]*accuracyMax[tmp_block_binary_reverse]
            totalaccuracyCP+=frequencyRevBiBlock[tmp_block_binary_reverse]*accuracyCP[tmp_block_binary_reverse]
            count=frequencyRevBiBlock[tmp_block_binary_reverse]
            OSACAMax=OSACAmaxCyclesRevBiBlock[tmp_block_binary_reverse]
            OSACACP=OSACACPCyclesRevBiBlock[tmp_block_binary_reverse]
            realBHive=float(BhiveCyclesRevBiBlock[tmp_block_binary_reverse])
            tmp=(OSACAMax+OSACACP)/2 * BHiveCount - realBHive
            totalOSACAavg+=abs(tmp)/realBHive * count
            fwriteblockfreq.writelines("{:5d} ".format(lineNum)+','+tmp_block_binary_reverse+','+str(frequencyRevBiBlock[tmp_block_binary_reverse])+','+ \
                str(OSACAmaxCyclesRevBiBlock[tmp_block_binary_reverse])+','+
                str(OSACACPCyclesRevBiBlock[tmp_block_binary_reverse])+','+
                str(llvmmcaCyclesRevBiBlock[tmp_block_binary_reverse])+','+
                str(BhiveCyclesRevBiBlock[tmp_block_binary_reverse])+','+
                str(accuracyLLVM[tmp_block_binary_reverse])+","+
                str(accuracyMax[tmp_block_binary_reverse])+','+
                str(accuracyCP[tmp_block_binary_reverse])+"\n")
        else:
            unvalidNum+=1
            fwriteblockfreq.writelines("{:5d} ".format(lineNum)+','+tmp_block_binary_reverse+','+str(frequencyRevBiBlock[tmp_block_binary_reverse])+','+ \
                str(OSACAmaxCyclesRevBiBlock[tmp_block_binary_reverse])+','+
                str(OSACACPCyclesRevBiBlock[tmp_block_binary_reverse])+','+
                str(llvmmcaCyclesRevBiBlock[tmp_block_binary_reverse])+','+
                str(BhiveCyclesRevBiBlock[tmp_block_binary_reverse])+','+
                str(accuracyLLVM[tmp_block_binary_reverse])+","+
                str(accuracyMax[tmp_block_binary_reverse])+','+
                str(accuracyCP[tmp_block_binary_reverse])+",wrong\n")
            wrongResultFile.writelines("{:5d} ".format(lineNum)+','+tmp_block_binary_reverse+','+str(frequencyRevBiBlock[tmp_block_binary_reverse])+','+ \
                str(OSACAmaxCyclesRevBiBlock[tmp_block_binary_reverse])+','+
                str(OSACACPCyclesRevBiBlock[tmp_block_binary_reverse])+','+
                str(llvmmcaCyclesRevBiBlock[tmp_block_binary_reverse])+','+
                str(BhiveCyclesRevBiBlock[tmp_block_binary_reverse])+','+
                str(accuracyLLVM[tmp_block_binary_reverse])+","+
                str(accuracyMax[tmp_block_binary_reverse])+','+
                str(accuracyCP[tmp_block_binary_reverse])+",wrong\n")
    fwriteblockfreq.writelines("validTotalNum & unvalidBlockNum :"+str(validNum)+" "+str(unvalidNum)+"\n")
    fwriteblockfreq.writelines("avg llvm error rate is "+str(totalAccuracyLLVM/validNum)+"\n")
    fwriteblockfreq.writelines("avg osaca Max error rate is "+str(totalaccuracyMax/validNum)+"\n")
    fwriteblockfreq.writelines("avg osaca CP error rate is "+str(totalaccuracyCP/validNum)+"\n")
    fwriteblockfreq.writelines("avg osaca avg error rate is "+str(totalOSACAavg/validNum)+"\n")
    print("avg LLVM error rate is "+str(totalAccuracyLLVM/validNum)+"\n") 
    print("avg OSACA Max error rate is "+str(totalaccuracyMax/validNum)+"\n") 
    print("avg OSACA CP error rate is "+str(totalaccuracyCP/validNum)+"\n") 
    print("avg osaca avg error rate is "+str(totalOSACAavg/validNum)+"\n") 
    fwriteblockfreq.close()
    wrongResultFile.close()

def checkFile(taskfilePath):
    tmpOSACAfilePath=taskfilePath+"/tmpOSACAfiles"
    mkdir(tmpOSACAfilePath)
    return tmpOSACAfilePath

def mkdir(path):
	folder = os.path.exists(path)
	if not folder:                   #???????????????????????????????????????????????????????????????
		os.makedirs(path)            #makedirs ?????????????????????????????????????????????????????????
		print("---  new folder...  ---")
	else:
		print("---  There is this folder!  ---")

def add2Excel(wb,name,isFirstSheet,unique_revBiblock,frequencyRevBiBlock,OSACAmaxCyclesRevBiBlock,OSACACPCyclesRevBiBlock,OSACALCDCyclesRevBiBlock,BhiveCyclesRevBiBlock,accuracyMax,accuracyCP,llvmmcaCyclesRevBiBlock,accuracyLLVM):
    # if isFirstSheet==1:
        # ws = wb.active # ?????????Sheet
        # ws.title = name
    wb.create_sheet(name)
    ws = wb[name]
    ws.append(["num","block_binary" , "ARM64_assembly_code", "block_frequency", "OSACA_max_result", "OSACA_CP","OSACA_LCD","LLVM-MCA_result", "BHive", "accuracyLLVM", "accuracyMax", "accuracyCP" ]) # ?????????
    ws.column_dimensions['B'].width = 62 # ????????????
    for i in ['C','D','E','F','G','H','I','J','K']:
        ws.column_dimensions[i].width = 15 # ????????????
    validNum=0
    unvalidNum=0
    totalAccuracyLLVM=0.0
    totalaccuracyMax=0.0
    totalaccuracyCP=0.0
    totalOSACAavg=0.0
    lineNum=0
    for tmp_block_binary_reverse in unique_revBiblock:
        lineNum+=1
        tmpInput=[]
        tmpARMassembly=capstone(capstoneInput(tmp_block_binary_reverse))
        if accuracyMax[tmp_block_binary_reverse] != 0 and accuracyCP[tmp_block_binary_reverse] != 0 and accuracyLLVM[tmp_block_binary_reverse] != 0:
            validNum+=frequencyRevBiBlock[tmp_block_binary_reverse]
            totalAccuracyLLVM+=frequencyRevBiBlock[tmp_block_binary_reverse]*accuracyLLVM[tmp_block_binary_reverse]
            totalaccuracyMax+=frequencyRevBiBlock[tmp_block_binary_reverse]*accuracyMax[tmp_block_binary_reverse]
            totalaccuracyCP+=frequencyRevBiBlock[tmp_block_binary_reverse]*accuracyCP[tmp_block_binary_reverse]
            count=frequencyRevBiBlock[tmp_block_binary_reverse]
            OSACAMax=OSACAmaxCyclesRevBiBlock[tmp_block_binary_reverse]
            OSACACP=OSACACPCyclesRevBiBlock[tmp_block_binary_reverse]
            realBHive=float(BhiveCyclesRevBiBlock[tmp_block_binary_reverse])
            tmp=(OSACAMax+OSACACP)/2 * BHiveCount - realBHive
            totalOSACAavg+=abs(tmp)/realBHive * count
            ws.append(["{:5d} ".format(lineNum), 
                tmp_block_binary_reverse,
                tmpARMassembly,
                frequencyRevBiBlock[tmp_block_binary_reverse],
                OSACAmaxCyclesRevBiBlock[tmp_block_binary_reverse],
                OSACACPCyclesRevBiBlock[tmp_block_binary_reverse],
                OSACALCDCyclesRevBiBlock[tmp_block_binary_reverse],
                llvmmcaCyclesRevBiBlock[tmp_block_binary_reverse],
                BhiveCyclesRevBiBlock[tmp_block_binary_reverse],
                accuracyLLVM[tmp_block_binary_reverse],
                accuracyMax[tmp_block_binary_reverse],
                accuracyCP[tmp_block_binary_reverse]])
        else:
            unvalidNum+=1
            ws.append(["{:5d} ".format(lineNum), 
                tmp_block_binary_reverse,
                tmpARMassembly,
                frequencyRevBiBlock[tmp_block_binary_reverse],
                OSACAmaxCyclesRevBiBlock[tmp_block_binary_reverse],
                OSACACPCyclesRevBiBlock[tmp_block_binary_reverse],
                OSACALCDCyclesRevBiBlock[tmp_block_binary_reverse],
                llvmmcaCyclesRevBiBlock[tmp_block_binary_reverse],
                BhiveCyclesRevBiBlock[tmp_block_binary_reverse],
                accuracyLLVM[tmp_block_binary_reverse],
                accuracyMax[tmp_block_binary_reverse],
                accuracyCP[tmp_block_binary_reverse],
                "ops!"])
    llvmerror=totalAccuracyLLVM/validNum
    osacaerror=totalOSACAavg/validNum
    return [llvmerror,osacaerror]

def excelGraphInit(wb):
    ws = wb.active # ?????????Sheet
    ws.title = 'Graph'
    ws = wb["Graph"]
    ws.column_dimensions['A'].width = 15 # ????????????
    ws.column_dimensions['B'].width = 15 # ????????????
    ws.column_dimensions['C'].width = 15 # ????????????
    ws.append(["applications","LLVM-MCA_error","OSACA_error"])

def excelGraphAdd(wb,taskName,llvmerror,osacaerror):
    ws = wb["Graph"]
    ws.append([taskName,llvmerror,osacaerror])

def excelGraphBuild(wb):
    # ??????????????????
    ws = wb["Graph"]
    ct_bar = BarChart()
    taskNum=len(taskList)
    ws['D1'] = '????????????'
    for i in range(2, taskNum+2):
        ws[f'D{i}'] = ws[f'C{i}'].value / ws[f'B{i}'].value
    d_ref = Reference(ws, min_col=2, min_row=1, max_row=taskNum+1, max_col=3)
    ct_bar.add_data(d_ref, titles_from_data=True)
    series = Reference(ws, min_col=1, min_row=2, max_row=taskNum+1)
    ct_bar.set_categories(series)
    ct_bar.x_axis.title = '??????'
    ct_bar.y_axis.title = '??????'
    ct_bar.y_axis.majorGridlines = None
    ct_bar.title = '????????????????????????????????????'
    ct_line = LineChart()
    d_ref = Reference(ws, min_col=4, min_row=1, max_row=taskNum+1)
    ct_line.add_data(d_ref, titles_from_data=True)
    ct_line.y_axis.axId = 200 # ???????????????
    ct_line.y_axis.title = '????????????'
    # ???????????????????????????????????????
    ct_line.y_axis.crosses = 'max'
    ct_bar += ct_line # ?????????+=?????????????????????+
    ws.add_chart(ct_bar, 'A10')
    wb.save(excelOutPath)

if __name__ == "__main__":
    global filename,taskfilenameprefixWithoutPath,taskfilenameprefix
    # print("?????????sudo??????")
    password="acsa1411"
    checkFile(taskfilePath)
    wb = Workbook()
    excelGraphInit(wb)
    isFirstSheet=1
    for taskKey, taskName in taskList.items():
        taskfilenameprefixWithoutPath=taskKey
        taskfilenameprefix="{}/{}".format(taskfilePath,taskfilenameprefixWithoutPath)
        taskfilenamesubfix="log"
        filename="{}.{}".format(taskfilenameprefix,taskfilenamesubfix)

        unique_revBiblock=set()
        frequencyRevBiBlock = defaultdict(int)
        llvmmcaCyclesRevBiBlock = defaultdict(int)
        OSACAmaxCyclesRevBiBlock = defaultdict(int)
        OSACACPCyclesRevBiBlock = defaultdict(int)
        OSACALCDCyclesRevBiBlock = defaultdict(int)
        BhiveCyclesRevBiBlock = defaultdict(int)
        accuracyLLVM = defaultdict(float)
        accuracyMax = defaultdict(float)
        accuracyCP = defaultdict(float)

        unique_revBiblock=readPartFile(password, unique_revBiblock,frequencyRevBiBlock,OSACAmaxCyclesRevBiBlock,OSACACPCyclesRevBiBlock,OSACALCDCyclesRevBiBlock,BhiveCyclesRevBiBlock,accuracyMax,accuracyCP,llvmmcaCyclesRevBiBlock,accuracyLLVM)
        print("blockSize {} {}".format(len(unique_revBiblock),len(frequencyRevBiBlock)))
        saveAllResult(taskfilenameprefix,unique_revBiblock,frequencyRevBiBlock,OSACAmaxCyclesRevBiBlock,OSACACPCyclesRevBiBlock,BhiveCyclesRevBiBlock,accuracyMax,accuracyCP,llvmmcaCyclesRevBiBlock,accuracyLLVM)
        [llvmerror,osacaerror] = add2Excel(wb,taskName,isFirstSheet,unique_revBiblock,frequencyRevBiBlock,OSACAmaxCyclesRevBiBlock,OSACACPCyclesRevBiBlock,OSACALCDCyclesRevBiBlock,BhiveCyclesRevBiBlock,accuracyMax,accuracyCP,llvmmcaCyclesRevBiBlock,accuracyLLVM)
        # add2Excel(wb,taskName,isFirstSheet,unique_revBiblock,frequencyRevBiBlock,OSACAmaxCyclesRevBiBlock,OSACACPCyclesRevBiBlock,OSACALCDCyclesRevBiBlock,BhiveCyclesRevBiBlock,accuracyMax,accuracyCP,llvmmcaCyclesRevBiBlock,accuracyLLVM)
        excelGraphAdd(wb,taskName,llvmerror,osacaerror)
        isFirstSheet=0
    excelGraphBuild(wb)
    # wb.save(excelOutPath)
