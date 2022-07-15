from icecream import ic
from collections import defaultdict

HistoryDataFile = "/home/shaojiemike/blockFrequency/Summary_BHiveCount5002022-07-15-16-15-36_tsj.xlsx"
pageName="tensorflow_2"
pageNameNotExisted="yahaha"

# https://blog.csdn.net/David_Dai_1108/article/details/78702032
def readExcelSet(pageName, columnName):
    unique_revBiblock=set()
    if not columnName=="block_binary":
        return unique_revBiblock
    from openpyxl import load_workbook
    # 只读模式打开文件
    wb = load_workbook(HistoryDataFile, read_only=True)
    # 获得所有 sheet 的名称()
    name_list = wb.sheetnames
    # 根据 sheet 名字获得 sheet
    if pageName not in name_list:
        ic(pageName,"Not existed")
        return unique_revBiblock
    ic(pageName,"existed")
    my_sheet = wb[pageName]
    for rowNum in range(2,my_sheet.max_row):
        ic(rowNum)
        unique_revBiblock.add(my_sheet["B"+str(rowNum)].value)
    return unique_revBiblock

def readExcelDict(pageName,columnName):
    readDefaultdict = defaultdict(int)
    if not(columnName=="BHive" or columnName=="Baseline"):
        return readDefaultdict
    elif columnName=="BHive":
        columnLetter="H"
    elif columnName=="Baseline":
        columnLetter="G"
    from openpyxl import load_workbook
    # 只读模式打开文件
    wb = load_workbook(HistoryDataFile, read_only=True)
    # 获得所有 sheet 的名称()
    name_list = wb.sheetnames
    # 根据 sheet 名字获得 sheet
    if pageName not in name_list:
        ic(pageName,"Not existed")
        return readDefaultdict
    ic(pageName,"existed")
    my_sheet = wb[pageName]
    for rowNum in range(2,my_sheet.max_row):
        ic(rowNum)
        readDefaultdict[my_sheet["B"+str(rowNum)].value]=int(my_sheet[columnLetter+str(rowNum)].value)
    return readDefaultdict

if ic(readExcelSet(pageName,"block_binary")):
    ic("Not Empty")
if not ic(readExcelSet(pageNameNotExisted,"block_binary")):
    ic("Empty")
ic(readExcelSet(pageNameNotExisted,"block_binary2e2e"))
ic(readExcelDict(pageName,"BHive"))
ic(readExcelDict(pageName,"Baseline"))

